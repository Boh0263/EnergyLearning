import json
import os
import pandas as pd
from pymongo import MongoClient
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


MONGO_USER = "rootarded"
MONGO_PASSWORD = "<Password>"
MONGO_HOST = ""
MONGO_PORT = "27017"
DB_NAME = "EnergyLearning"
DB_NAME_LOCAL_h = "EnergyLearningLocal"
MONGO_URI_LOCAL_h = f"mongodb://localhost:{MONGO_PORT}/"
MONGO_URI = f"mongodb://{MONGO_USER}:{MONGO_PASSWORD}@{MONGO_HOST}:{MONGO_PORT}/{DB_NAME}"

JSON_FILE_PATH = "Groups.json"
SAVE_DIRECTORY = "..\Data\Results"
FIELDS_TO_EXTRACT = ["weightedAvgTotalCo2Consumption", "weightedAvgTotalConsumption", "avgRenewableRatio", "avgFossilFuelRatio"]

def load_groups_json():
    try:
        with open(JSON_FILE_PATH, "r") as file:
            groups = json.load(file)
        return groups
    except Exception as e:
        print(f"Error loading JSON file: {e}")
        return None

def get_color_scale(value, min_value, max_value):
    if value is None:
        return PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    if max_value != min_value:
        normalized_value = (value - min_value) / (max_value - min_value)
    else:
        normalized_value = 0.5
    if normalized_value < 0.33:
        color = "FF0000"
    elif normalized_value < 0.66:
        color = "FFFF00"
    else:
        color = "00FF00"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")

def format_excel(ws):
    title = "Energy Consumption & CO2 Data"
    ws.insert_rows(1)
    ws.merge_cells("A1:E1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    header_fill = PatternFill(start_color="B0C4DE", fill_type="solid")
    bold_font = Font(bold=True)
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    for cell in ws[2]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border_style
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border_style
    for col in ws.iter_cols(min_row=2, max_row=ws.max_row):
        max_length = 0
        col_letter = col[0].column_letter if col[0].column else None
        if col_letter:
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2


def add_totals(ws):
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)

    col_letter_2 = openpyxl.utils.get_column_letter(2)
    sum_formula_2 = f"=SUM({col_letter_2}2:{col_letter_2}{total_row - 1})"
    ws.cell(row=total_row, column=2, value=sum_formula_2).font = Font(bold=True)

    col_letter_3 = openpyxl.utils.get_column_letter(3)
    sum_formula_3 = f"=SUM({col_letter_3}2:{col_letter_3}{total_row - 1})"
    ws.cell(row=total_row, column=3, value=sum_formula_3).font = Font(bold=True)

    col_letter_4 = openpyxl.utils.get_column_letter(4)
    sumproduct_formula_4 = f"=SUMPRODUCT({col_letter_4}2:{col_letter_4}{total_row - 1},{col_letter_2}2:{col_letter_2}{total_row - 1}) / SUM({col_letter_2}2:{col_letter_2}{total_row - 1})"
    ws.cell(row=total_row, column=4, value=sumproduct_formula_4).font = Font(bold=True)

    col_letter_5 = openpyxl.utils.get_column_letter(5)
    sumproduct_formula_5 = f"=SUMPRODUCT({col_letter_5}2:{col_letter_5}{total_row - 1},{col_letter_2}2:{col_letter_2}{total_row - 1}) / SUM({col_letter_2}2:{col_letter_2}{total_row - 1})"
    ws.cell(row=total_row, column=5, value=sumproduct_formula_5).font = Font(bold=True)


def find_data():
    client = MongoClient(MONGO_URI_LOCAL_h)
    db = client[DB_NAME_LOCAL_h]
    collection = db["weightedZoneResultsv2"]
    groups = load_groups_json()
    if not groups:
        print("No groups loaded. Exiting.")
        return
    for idx, group in enumerate(groups):
        data_dict = {field: [] for field in FIELDS_TO_EXTRACT}
        for key in group:
            document = collection.find_one({"_id": key}, {field: 1 for field in FIELDS_TO_EXTRACT})
            if document:
                for field in FIELDS_TO_EXTRACT:
                    data_dict[field].append(document.get(field, None))
            else:
                for field in FIELDS_TO_EXTRACT:
                    data_dict[field].append(None)
        df = pd.DataFrame(data_dict)
        df["ID"] = group
        df = df.set_index("ID")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        for col_num, column_name in enumerate(df.columns, 2):
            ws.cell(row=1, column=col_num, value=column_name)
        for row_idx, param in enumerate(df.index, 2):
            ws.cell(row=row_idx, column=1, value=param)
            for col_idx, (col_name, value) in enumerate(df.loc[param].items(), 2):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = get_color_scale(value, df[col_name].min() if df[col_name].notna().any() else 0, df[col_name].max() if df[col_name].notna().any() else 1)
        format_excel(ws)
        add_totals(ws)
        output_file_name = f"group_{idx + 1}.xlsx"
        file_path = os.path.join(SAVE_DIRECTORY, output_file_name)
        wb.save(file_path)
        print(f"Saved: {output_file_name}")

if __name__ == "__main__":
    find_data()
